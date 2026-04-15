import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def create_tz_xlsx(filename="ТЗ_Оверлей_Печать_v2.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ТЗ_Оверлей"

    # === СТИЛИ ===
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(name="Calibri", bold=True, size=11, color="2F5496")
    normal_font = Font(name="Calibri", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def style_cell(row, col, value, font=normal_font, fill=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        cell.alignment = wrap_align
        cell.border = thin_border
        if fill: cell.fill = fill
        return cell

    # === ЗАГОЛОВОК ===
    ws.merge_cells("A1:C1")
    ws["A1"] = "ТЕХНИЧЕСКОЕ ЗАДАНИЕ: Модуль оверлей-печати для парка ТС"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # === ШАПКА ТАБЛИЦЫ ===
    headers = ["Раздел", "Параметр / Поле", "Описание и требования"]
    for i, h in enumerate(headers, 1):
        style_cell(3, i, h, header_font, header_fill)
    ws.row_dimensions[3].height = 20

    # === ДАННЫЕ ===
    data = [
        ["1. Общее назначение", "Наименование проекта", "Модуль оверлей-печати, интегрированный в веб-калькулятор ГСМ"],
        ["", "Цель", "Бесшовный переход от расчёта к печати путевых листов для парка ТС разной комплектации"],
        ["", "Архитектура", "Единый HTML-файл. Поддержка мульти-профилей ТС и доп. оборудования. Офлайн-рендеринг"],
        
        ["2. Рабочий процесс", "Шаг 1: Выбор профиля", "Водитель/диспетчер выбирает ТС из выпадающего списка. Автоматическая подгрузка норм и карты наложения"],
        ["", "Шаг 2: Расчёт", "Калькулятор учитывает профиль ТС, считает остаток, пробег, нормы. Данные передаются в модуль печати"],
        ["", "Шаг 3: Доп. данные", "Динамический показ полей в зависимости от типа ТС/флага доп. оборудования"],
        ["", "Шаг 4: Печать", "Генерация PDF/PNG → отправка на Pantum через мобильное приложение"],
        
        ["3. Конфигурация ТС", "Профиль ТС", "JSON-файл с нормами расхода, списком обязательных полей и привязкой к кодам ячеек"],
        ["", "Доп. оборудование", "Флаг has_extra_eq. При активации добавляются специфичные строки/коды наложения"],
        
        ["4. Карта наложения", "Система кодировки", "Короткие 3-символьные коды: Л01-Л99 (лицевая), О01-О99 (оборотная), Д01-Д99 (доп. техника). Оптимально для мелких ячеек"],
        ["", "Формат карты", '{"Л01": {"x": 15.0, "y": 42.0, "w": 12, "h": 4, "font": "8pt"}, "О01": {"x": 15.0, "y": 42.0}, "Д01": {"x": 85.0, "y": 120.0}}'],
        ["", "Логика дублирования", '{"номер_путевки": ["Л01", "О01"], "остаток": ["Л02", "Д01"]}. Изменение в одном поле обновляет все связанные коды до рендера'],
        
        ["5. Интеграция в калькулятор", "UI/UX", "Вкладка «Печать» с выбором ТС. Сохранение текущего дизайна, тёмной темы и адаптивности"],
        ["", "Передача данных", "Чтение из JS-состояния калькулятора. Без localStorage и внешних API"],
        ["", "Офлайн/Производительность", "≤ 2 сек генерация на мобильном. WebWorker/Canvas. Полная автономность"],
        
        ["6. Требования к печати", "Режим", "Duplex (двусторонняя). Чёткое позиционирование оверлея относительно типографских ячеек"],
        ["", "Оборудование", "Pantum App + принтер. PDF/PNG, фиксированный DPI, масштаб 100%"],
        ["", "Точность", "Допустимое смещение ≤ 0.5 мм. Встроенный предпросмотр + режим калибровки"],
        
        ["7. Технические требования", "Конфигурация", "Внешние JSON: vehicle_profiles.json, overlay_map.json. Динамическая подгрузка без перезагрузки"],
        ["", "Расширяемость", "Добавление ТС/полей/кодов через правку JSON. Архитектура поддерживает N профилей"],
        ["", "Безопасность", "Все вычисления локально. Данные не передаются во внешние сервисы, не кэшируются"]
    ]

    # === ЗАПИСЬ И ФОРМАТИРОВАНИЕ ===
    current_row = 4
    for row_data in data:  # ✅ Исправлено: добавлено "data:"
        section, param, desc = row_data
        r = current_row
        
        cell_sec = ws.cell(row=r, column=1)
        cell_sec.value = section
        cell_sec.font = section_font if section else Font(name="Calibri", size=11, color="808080")
        cell_sec.alignment = wrap_align
        cell_sec.border = thin_border
        if section: cell_sec.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        style_cell(r, 2, param)
        style_cell(r, 3, desc)
        ws.row_dimensions[r].height = 35
        current_row += 1

    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 90

    # Сохранение
    wb.save(filename)
    print(f"✅ Файл успешно создан: {filename}")

if __name__ == "__main__":
    create_tz_xlsx()